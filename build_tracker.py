import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, sys

# ── Config ────────────────────────────────────────────────────────────────────
TODAY = "3/20/26"

XLSX_PATH = "/Users/michaelwhitney/Documents/AI STUFF/HTML files created by Claudette/Bambu_Filament_Tracker.xlsx"

# Colors: list of (name, inStock) tuples — updated each run by Claude
FILAMENTS = {
    "PLA Basic": {
        "url": "https://us.store.bambulab.com/products/pla-basic-filament",
        "colors": [
            ("Jade White",      True),
            ("Orange",          True),
            ("Blue",            False),
            ("Gray",            True),
            ("Beige",           True),
            ("Silver",          True),
            ("Yellow",          True),
            ("Blue Grey",       False),
            ("Pink",            False),
            ("Brown",           True),
            ("Red",             True),
            ("Black",           True),
            ("Bambu Green",     True),
            ("Bronze",          True),
            ("Gold",            True),
            ("Purple",          False),
            ("Magenta",         False),
            ("Cyan",            True),
            ("Mistletoe Green", True),
            ("Light Gray",      True),
            ("Dark Gray",       False),
            ("Maroon Red",      True),
            ("Sunflower Yellow",True),
            ("Turquoise",       True),
            ("Indigo Purple",   False),
            ("Bright Green",    False),
            ("Cocoa Brown",     True),
            ("Hot Pink",        False),
            ("Cobalt Blue",     False),
            ("Pumpkin Orange",  True),
        ],
    },
    "PLA Matte": {
        "url": "https://us.store.bambulab.com/products/pla-matte",
        "colors": [
            ("Matte Ivory White",    True),
            ("Matte Ash Gray",       True),
            ("Matte Mandarin Orange",True),
            ("Matte Sakura Pink",    False),
            ("Matte Lemon Yellow",   True),
            ("Matte Charcoal",       True),
            ("Matte Grass Green",    True),
            ("Matte Latte Brown",    True),
            ("Matte Scarlet Red",    True),
            ("Matte Ice Blue",       False),
            ("Matte Lilac Purple",   True),
            ("Matte Marine Blue",    True),
            ("Matte Dark Red",       True),
            ("Matte Dark Blue",      False),
            ("Matte Dark Brown",     True),
            ("Matte Dark Green",     True),
            ("Matte Desert Tan",     False),
            ("Matte Bone White",     False),
            ("Matte Plum",           False),
            ("Matte Sky Blue",       False),
            ("Matte Apple Green",    True),
            ("Matte Dark Chocolate", True),
            ("Matte Caramel",        False),
            ("Matte Terracotta",     False),
            ("Matte Nardo Gray",     True),
        ],
    },
    "PETG Basic": {
        "url": "https://us.store.bambulab.com/products/petg-basic",
        "colors": [
            ("Black",      True),
            ("White",      True),
            ("Gray",       True),
            ("Red",        True),
            ("Yellow",     True),
            ("Reflex Blue",True),
            ("Dark Brown", True),
        ],
    },
    "PETG HF": {
        "url": "https://us.store.bambulab.com/products/petg-hf",
        "colors": [
            ("Blue",         False),
            ("Red",          True),
            ("Black",        True),
            ("Gray",         False),
            ("White",        False),
            ("Dark Gray",    False),
            ("Cream",        False),
            ("Orange",       True),
            ("Green",        True),
            ("Yellow",       False),
            ("Lime Green",   False),
            ("Forest Green", False),
            ("Lake Blue",    False),
            ("Peanut Brown", False),
        ],
    },
    "TPU 95A HF": {
        "url": "https://us.store.bambulab.com/products/tpu-95a-hf",
        "colors": [
            ("Black",  True),
            ("White",  False),
            ("Gray",   True),
            ("Yellow", True),
            ("Blue",   True),
            ("Red",    False),
        ],
    },
    "TPU for AMS": {
        "url": "https://us.store.bambulab.com/products/tpu-for-ams",
        "colors": [
            ("Red",       True),
            ("Yellow",    False),
            ("Blue",      True),
            ("Neon Green",True),
            ("White",     True),
            ("Gray",      True),
            ("Black",     True),
        ],
    },
    "TPU 85A-90A": {
        "url": "https://us.store.bambulab.com/products/tpu-85a-tpu-90a",
        "colors": [
            ("85A — Light Cyan",   True),
            ("85A — Neon Orange",  True),
            ("85A — Black",        True),
            ("85A — Flesh",        True),
            ("85A — Lime Green",   True),
            ("90A — Black",        True),
            ("90A — White",        True),
            ("90A — Quicksilver",  True),
            ("90A — Crystal Blue", True),
            ("90A — Grape Jelly",  True),
            ("90A — Cocoa Brown",  True),
            ("90A — Frozen",       True),
            ("90A — Blaze",        True),
        ],
    },
}

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(color="FFFFFF", bold=False, size=10):
    return Font(color=color, bold=bold, size=size)

def align(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v)

thin = Side(style="thin", color="333344")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

FILL_TITLE   = fill("1A1A2E")
FILL_HEADER  = fill("0F3460")
FILL_ROW1    = fill("1E1E2E")
FILL_ROW2    = fill("252535")
FILL_INSTOCK = fill("1B4332")
FILL_OOS     = fill("5C1C1C")

def apply(cell, value=None, fll=None, fnt=None, aln=None, brd=None):
    if value is not None:
        cell.value = value
    if fll:  cell.fill = fll
    if fnt:  cell.font = fnt
    if aln:  cell.alignment = aln
    if brd:  cell.border = brd

# ── Load or create workbook ───────────────────────────────────────────────────
if os.path.exists(XLSX_PATH):
    wb = openpyxl.load_workbook(XLSX_PATH)
    print(f"Loaded existing workbook: {XLSX_PATH}")
    append_mode = True
else:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    print("Creating new workbook")
    append_mode = False

# ── Per-filament sheets ───────────────────────────────────────────────────────
for name, info in FILAMENTS.items():
    colors = info["colors"]

    if append_mode and name in wb.sheetnames:
        ws = wb[name]
        # Check if TODAY already exists as a column header (avoid duplicates)
        existing_dates = [ws.cell(row=2, column=c).value for c in range(3, ws.max_column + 1)]
        if TODAY in existing_dates:
            print(f"  {name}: skipping — {TODAY} already exists")
            continue
        # Find first empty column slot starting at col 3 (fills gaps from pre-formatting)
        next_col = 3
        while ws.cell(row=2, column=next_col).value is not None:
            next_col += 1
        # Add date header
        c = ws.cell(row=2, column=next_col)
        apply(c, TODAY, FILL_HEADER, font("FFFFFF", bold=True), align("center"), BORDER)
        # Add status for each color row
        for i, (color, in_stock) in enumerate(colors):
            row = i + 3
            status = "✅" if in_stock else "🚫"
            row_fill = FILL_ROW1 if i % 2 == 0 else FILL_ROW2
            status_fill = FILL_INSTOCK if in_stock else FILL_OOS
            apply(ws.cell(row=row, column=next_col), status, status_fill, font("FFFFFF"), align("center"), BORDER)
        ws.column_dimensions[get_column_letter(next_col)].width = 12
        print(f"  {name}: appended column {get_column_letter(next_col)} ({TODAY})")
    else:
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 12
        for col in range(4, 14):
            ws.column_dimensions[get_column_letter(col)].width = 12
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 22
        # Title row
        ws.merge_cells("A1:C1")
        apply(ws["A1"], f"{name} — Availability Tracker",
              FILL_TITLE, font("FFFFFF", bold=True, size=11), align("left"))
        apply(ws["D1"], info["url"], FILL_TITLE, font("888888"), align("left"))
        # Header row
        for col, h in enumerate(["#", "Color", TODAY], 1):
            c = ws.cell(row=2, column=col)
            apply(c, h, FILL_HEADER, font("FFFFFF", bold=True), align("center"), BORDER)
        # Color rows
        for i, (color, in_stock) in enumerate(colors):
            row = i + 3
            row_fill = FILL_ROW1 if i % 2 == 0 else FILL_ROW2
            status = "✅" if in_stock else "🚫"
            status_fill = FILL_INSTOCK if in_stock else FILL_OOS
            apply(ws.cell(row=row, column=1), i + 1,  row_fill, font("888888"), align("center"), BORDER)
            apply(ws.cell(row=row, column=2), color,   row_fill, font("FFFFFF"), align("left"),   BORDER)
            apply(ws.cell(row=row, column=3), status,  status_fill, font("FFFFFF"), align("center"), BORDER)
        ws.freeze_panes = "A3"
        print(f"  {name}: created sheet with {len(colors)} colors")

# ── Summary sheet ─────────────────────────────────────────────────────────────
if "Summary" in wb.sheetnames:
    del wb["Summary"]

# Insert Summary at position 0
ws = wb.create_sheet("Summary", 0)

ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 11
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 52
ws.row_dimensions[1].height = 26
ws.row_dimensions[2].height = 22

# Title row
ws.merge_cells("A1:E1")
apply(ws["A1"], "Bambu Lab Filament Availability — Summary",
      FILL_TITLE, font("FFFFFF", bold=True, size=12), align("left"))
apply(ws["F1"], f"Last updated: {TODAY}",
      FILL_TITLE, font("AAAAAA"), align("right"))

# Header row
for col, h in enumerate(["Filament Line","Total Colors","In Stock","Out of Stock","% In Stock","Store Link"], 1):
    c = ws.cell(row=2, column=col)
    apply(c, h, FILL_HEADER, font("FFFFFF", bold=True), align("center"), BORDER)

# Data rows
for i, (name, info) in enumerate(FILAMENTS.items()):
    row = i + 3
    row_fill = FILL_ROW1 if i % 2 == 0 else FILL_ROW2
    colors = info["colors"]
    total = len(colors)
    in_stk = sum(1 for _, s in colors if s)
    out_stk = total - in_stk
    pct = f"{round(in_stk/total*100)}%"

    apply(ws.cell(row=row, column=1), name,       row_fill, font("FFFFFF"), align("left"),   BORDER)
    apply(ws.cell(row=row, column=2), total,       row_fill, font("FFFFFF"), align("center"), BORDER)
    apply(ws.cell(row=row, column=3), in_stk,      row_fill, font("90EE90", bold=True), align("center"), BORDER)
    apply(ws.cell(row=row, column=4), out_stk if out_stk else 0,
          row_fill, font("FF6B6B", bold=True) if out_stk else font("FFFFFF"), align("center"), BORDER)
    apply(ws.cell(row=row, column=5), pct,         row_fill, font("FFFFFF"), align("center"), BORDER)
    apply(ws.cell(row=row, column=6), info["url"], row_fill, font("FFFFFF"), align("left"),   BORDER)

ws.freeze_panes = "A3"

# ── Save ──────────────────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(XLSX_PATH), exist_ok=True)
wb.save(XLSX_PATH)
print(f"Saved: {XLSX_PATH}")
for s in wb.sheetnames:
    print(f"  {s}: {wb[s].max_row - 2} rows")
