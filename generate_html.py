"""
Reads Bambu_Filament_Tracker.xlsx and generates index.html.
All historical date columns are embedded as JSON so a date picker
lets visitors browse past snapshots without a server.
"""

import json
import os
import openpyxl

XLSX_PATH = "/Users/michaelwhitney/Documents/AI STUFF/HTML files created by Claudette/Bambu_Filament_Tracker.xlsx"
OUT_PATH  = "/Users/michaelwhitney/Documents/3D PRINTING/bambu-filament-tracker/index.html"

SHEET_META = {
    "PLA Basic":   {"group": "PLA",  "url": "https://us.store.bambulab.com/products/pla-basic-filament"},
    "PLA Matte":   {"group": "PLA",  "url": "https://us.store.bambulab.com/products/pla-matte"},
    "PETG Basic":  {"group": "PETG", "url": "https://us.store.bambulab.com/products/petg-basic"},
    "PETG HF":     {"group": "PETG", "url": "https://us.store.bambulab.com/products/petg-hf"},
    "TPU 95A HF":  {"group": "TPU",  "url": "https://us.store.bambulab.com/products/tpu-95a-hf"},
    "TPU for AMS": {"group": "TPU",  "url": "https://us.store.bambulab.com/products/tpu-for-ams"},
    "TPU 85A-90A": {"group": "TPU",  "url": "https://us.store.bambulab.com/products/tpu-85a-tpu-90a"},
}
GROUPS = ["PLA", "PETG", "TPU"]


def read_spreadsheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # Dates live in the header row (row 2) of any filament sheet, columns C+
    first_ws = wb["PLA Basic"]
    header = list(first_ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    dates = [str(h) for h in header[2:] if h is not None]

    filaments = {}
    for sheet_name, meta in SHEET_META.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        colors = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[1]:
                continue
            name = str(row[1])
            statuses = []
            for i in range(len(dates)):
                v = row[2 + i] if (2 + i) < len(row) else None
                statuses.append(1 if v == "✅" else 0)
            colors[name] = statuses
        filaments[sheet_name] = {
            "group": meta["group"],
            "url":   meta["url"],
            "colors": colors,
        }

    return dates, filaments


def build_html(dates, filaments):
    payload = json.dumps({"dates": dates, "filaments": filaments}, separators=(",", ":"))

    return f"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Bambu Lab Filament Availability</title>
  <style>
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

    /* ── Tokens ── */
    :root {{
      --bg:          #111827;
      --surface:     #1f2937;
      --surface2:    #263244;
      --border:      #374151;
      --text:        #e5e7eb;
      --text-muted:  #9ca3af;
      --text-dim:    #6b7280;
      --in-bg:       #14532d;
      --in-fg:       #bbf7d0;
      --out-bg:      #450a0a;
      --out-fg:      #fca5a5;
      --in-stat:     #4ade80;
      --out-stat:    #f87171;
      --accent:      #3b82f6;
    }}
    [data-theme="light"] {{
      --bg:          #f3f4f6;
      --surface:     #ffffff;
      --surface2:    #f9fafb;
      --border:      #d1d5db;
      --text:        #111827;
      --text-muted:  #4b5563;
      --text-dim:    #9ca3af;
      --in-bg:       #dcfce7;
      --in-fg:       #166534;
      --out-bg:      #fee2e2;
      --out-fg:      #991b1b;
      --in-stat:     #16a34a;
      --out-stat:    #dc2626;
      --accent:      #2563eb;
    }}

    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: var(--bg);
      color: var(--text);
      min-height: 100vh;
      transition: background .2s, color .2s;
    }}

    /* ── Header ── */
    header {{
      background: var(--surface);
      border-bottom: 1px solid var(--border);
      padding: .85rem 1.25rem;
      display: flex;
      align-items: center;
      justify-content: space-between;
      flex-wrap: wrap;
      gap: .6rem;
      position: sticky;
      top: 0;
      z-index: 10;
    }}
    header h1 {{
      font-size: 1.05rem;
      font-weight: 700;
      color: var(--text);
      white-space: nowrap;
    }}
    .controls {{
      display: flex;
      align-items: center;
      gap: .6rem;
      flex-wrap: wrap;
    }}
    .date-label {{
      font-size: .78rem;
      color: var(--text-muted);
      white-space: nowrap;
    }}
    #date-picker {{
      background: var(--bg);
      color: var(--text);
      border: 1px solid var(--border);
      border-radius: 6px;
      padding: .3rem .55rem;
      font-size: .82rem;
      cursor: pointer;
    }}
    #theme-toggle {{
      background: var(--bg);
      border: 1px solid var(--border);
      border-radius: 6px;
      padding: .3rem .6rem;
      font-size: .9rem;
      cursor: pointer;
      color: var(--text);
      line-height: 1;
    }}
    #theme-toggle:hover {{ background: var(--surface2); }}

    /* ── Chart ── */
    .chart-wrap {{
      background: var(--surface);
      border-bottom: 1px solid var(--border);
      padding: 1rem 1.5rem .85rem;
    }}
    .chart-title {{
      font-size: .72rem;
      font-weight: 600;
      color: var(--text-dim);
      text-transform: uppercase;
      letter-spacing: .07em;
      margin-bottom: .6rem;
    }}

    /* ── Columns ── */
    .columns {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 1rem;
      padding: 1rem;
      max-width: 1200px;
      margin: 0 auto;
    }}

    /* ── Column card ── */
    .col-header {{
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 10px 10px 0 0;
      padding: .85rem 1rem .7rem;
      margin-bottom: 2px;
    }}
    .col-title {{
      font-size: 1.05rem;
      font-weight: 700;
      color: var(--text);
      margin-bottom: .3rem;
    }}
    .col-summary {{
      display: flex;
      gap: .7rem;
      font-size: .8rem;
      font-weight: 600;
      margin-bottom: .45rem;
    }}
    .prog-wrap {{
      height: 5px;
      background: var(--border);
      border-radius: 99px;
      overflow: hidden;
    }}
    .prog-bar {{
      height: 100%;
      border-radius: 99px;
      transition: width .3s;
    }}

    /* ── Collapsible sections ── */
    details {{
      background: var(--surface);
      border: 1px solid var(--border);
      border-top: none;
    }}
    details:last-child {{ border-radius: 0 0 10px 10px; }}
    summary {{
      display: flex;
      align-items: center;
      gap: .5rem;
      padding: .55rem 1rem;
      cursor: pointer;
      user-select: none;
      list-style: none;
    }}
    summary::-webkit-details-marker {{ display: none; }}
    .triangle {{
      font-size: 1.1rem;
      color: var(--text-dim);
      transition: transform .2s;
      flex-shrink: 0;
      display: inline-block;
      line-height: 1;
    }}
    details[open] .triangle {{ transform: rotate(90deg); }}
    summary:hover {{ background: var(--surface2); }}

    .section-name {{
      font-size: .88rem;
      font-weight: 600;
      color: var(--text);
      flex: 1;
    }}
    .section-meta {{
      display: flex;
      align-items: center;
      gap: .45rem;
      font-size: .75rem;
      font-weight: 600;
    }}
    .shop-link {{
      color: var(--text-dim);
      text-decoration: none;
      font-weight: 400;
      font-size: .72rem;
    }}
    .shop-link:hover {{ color: var(--text-muted); }}

    /* ── Stat colors ── */
    .si  {{ color: var(--in-stat); }}
    .so  {{ color: var(--out-stat); }}
    .pct {{ color: var(--text-muted); }}

    /* ── Pills ── */
    .pills {{
      display: flex;
      flex-wrap: wrap;
      gap: .28rem;
      padding: .55rem 1rem .7rem;
    }}
    .pill {{
      font-size: .7rem;
      padding: .18rem .42rem;
      border-radius: 99px;
      white-space: nowrap;
    }}
    .pill.in  {{ background: var(--in-bg);  color: var(--in-fg); }}
    .pill.out {{ background: var(--out-bg); color: var(--out-fg); }}

    /* ── Footer ── */
    footer {{
      text-align: center;
      padding: 2rem 1rem;
      font-size: .75rem;
      color: var(--text-dim);
      display: flex;
      justify-content: center;
      gap: 1.5rem;
      flex-wrap: wrap;
    }}
    footer a {{ color: var(--text-muted); text-decoration: none; }}
    footer a:hover {{ color: var(--text); }}

    /* ── Mobile ── */
    @media (max-width: 700px) {{
      .columns {{ grid-template-columns: 1fr; padding: .75rem; }}
      header h1 {{ font-size: .95rem; }}
    }}
  </style>
</head>
<body>

<header>
  <h1>🧵 Bambu Lab Filament Availability</h1>
  <div class="controls">
    <span class="date-label">Showing:</span>
    <select id="date-picker" onchange="render(this.value)"></select>
    <button id="theme-toggle" onclick="toggleTheme()" title="Toggle light/dark mode">☀️</button>
  </div>
</header>

<div class="chart-wrap">
  <div class="chart-title">Colors by Filament Line</div>
  <div id="chart"></div>
</div>

<div class="columns" id="columns"></div>

<footer>
  <span>Data from <a href="https://us.store.bambulab.com" target="_blank">us.store.bambulab.com</a> · Updated daily</span>
  <a href="mailto:plazman888@icloud.com">Contact</a>
</footer>

<script>
const DATA = {payload};
const GROUPS = ["PLA","PETG","TPU"];
const FILAMENT_NAMES = Object.keys(DATA.filaments);

/* ── Theme ── */
function toggleTheme() {{
  const html = document.documentElement;
  const isDark = html.getAttribute("data-theme") === "dark";
  html.setAttribute("data-theme", isDark ? "light" : "dark");
  document.getElementById("theme-toggle").textContent = isDark ? "🌙" : "☀️";
  localStorage.setItem("theme", isDark ? "light" : "dark");
}}
(function() {{
  const saved = localStorage.getItem("theme");
  if (saved) {{
    document.documentElement.setAttribute("data-theme", saved);
    document.getElementById("theme-toggle").textContent = saved === "light" ? "🌙" : "☀️";
  }}
}})();

/* ── Date picker ── */
function populatePicker() {{
  const sel = document.getElementById("date-picker");
  const dates = DATA.dates;
  dates.slice().reverse().forEach((d, i) => {{
    const opt = document.createElement("option");
    opt.value = d;
    opt.textContent = d;
    if (i === 0) opt.selected = true;
    sel.appendChild(opt);
  }});
}}

/* ── Stats helpers ── */
function statsForDate(colors, dateIdx) {{
  const vals = Object.values(colors).map(arr => arr[dateIdx] ?? 0);
  const total = vals.length;
  const inStk = vals.filter(v => v === 1).length;
  return {{ total, inStk, out: total - inStk, pct: total ? Math.round(inStk/total*100) : 0 }};
}}

function barColor(pct) {{
  return pct >= 75 ? "#4caf50" : pct >= 40 ? "#ff9800" : "#f44336";
}}

/* ── Chart ── */
function buildChart(dateIdx) {{
  const CHART_H = 140, BAR_W = 36, GAP = 18, PAD = 8, LABEL_H = 36;
  const SVG_H = CHART_H + LABEL_H + 10;
  const names = FILAMENT_NAMES;
  const maxTotal = Math.max(...names.map(n => Object.keys(DATA.filaments[n].colors).length));
  const SVG_W = PAD*2 + names.length*BAR_W + (names.length-1)*GAP;

  const shortLabels = ["PLA Basic","PLA Matte","PETG Basic","PETG HF","TPU 95A HF","TPU AMS","TPU 85/90A"];
  let bars = "", labels = "";

  names.forEach((name, i) => {{
    const {{total, inStk, out, pct}} = statsForDate(DATA.filaments[name].colors, dateIdx);
    const x = PAD + i*(BAR_W+GAP);
    const fullH = Math.round(total/maxTotal*CHART_H);
    const inH  = Math.round(inStk/total*fullH);
    const outH = fullH - inH;
    const yTop = CHART_H - fullH;

    if (outH > 0) bars += `<rect x="${{x}}" y="${{yTop}}" width="${{BAR_W}}" height="${{outH}}" fill="#c0504d" rx="3"/>`;
    if (inH  > 0) bars += `<rect x="${{x}}" y="${{yTop+outH}}" width="${{BAR_W}}" height="${{inH}}" fill="#4472c4" rx="3"/>`;

    const parts = (shortLabels[i]||name).split(" ");
    const mid = Math.ceil(parts.length/2);
    const line1 = parts.slice(0,mid).join(" ");
    const line2 = parts.slice(mid).join(" ");
    const lx = x + BAR_W/2;
    labels += `<text x="${{lx}}" y="${{CHART_H+14}}" text-anchor="middle" fill="var(--text-dim)" font-size="10" font-family="system-ui">${{line1}}</text>`;
    if (line2) labels += `<text x="${{lx}}" y="${{CHART_H+27}}" text-anchor="middle" fill="var(--text-dim)" font-size="10" font-family="system-ui">${{line2}}</text>`;
  }});

  const lx = SVG_W - 130;
  const legend = `
    <rect x="${{lx}}" y="4" width="12" height="12" fill="#4472c4" rx="2"/>
    <text x="${{lx+16}}" y="14" fill="var(--text-dim)" font-size="11" font-family="system-ui">In Stock</text>
    <rect x="${{lx+70}}" y="4" width="12" height="12" fill="#c0504d" rx="2"/>
    <text x="${{lx+86}}" y="14" fill="var(--text-dim)" font-size="11" font-family="system-ui">Out</text>`;

  document.getElementById("chart").innerHTML =
    `<svg viewBox="0 0 ${{SVG_W}} ${{SVG_H}}" width="100%" style="max-width:${{SVG_W}}px;display:block;margin:0 auto">
      ${{legend}}${{bars}}${{labels}}
    </svg>`;
}}

/* ── Columns ── */
function buildColumns(dateIdx) {{
  const wrap = document.getElementById("columns");
  wrap.innerHTML = "";

  GROUPS.forEach(group => {{
    const groupFilaments = FILAMENT_NAMES.filter(n => DATA.filaments[n].group === group);
    const allColors = groupFilaments.flatMap(n => Object.values(DATA.filaments[n].colors).map(arr => arr[dateIdx]??0));
    const total = allColors.length;
    const inStk = allColors.filter(v=>v===1).length;
    const out   = total - inStk;
    const pct   = total ? Math.round(inStk/total*100) : 0;
    const bc    = barColor(pct);

    // Sections
    let sectionsHTML = "";
    groupFilaments.forEach(name => {{
      const info = DATA.filaments[name];
      const s = statsForDate(info.colors, dateIdx);
      const pills = Object.entries(info.colors).map(([color, arr]) => {{
        const inS = (arr[dateIdx]??0) === 1;
        return `<span class="pill ${{inS?'in':'out'}}">${{inS?'✅':'🚫'}} ${{color}}</span>`;
      }}).join("");

      sectionsHTML += `
        <details open>
          <summary>
            <span class="triangle">&#9658;</span>
            <span class="section-name">${{name}}</span>
            <span class="section-meta">
              <span class="si">${{s.inStk}}✅</span>
              <span class="so">${{s.out}}🚫</span>
              <a class="shop-link" href="${{info.url}}" target="_blank" onclick="event.stopPropagation()">Shop ↗</a>
            </span>
          </summary>
          <div class="pills">${{pills}}</div>
        </details>`;
    }});

    const col = document.createElement("div");
    col.className = "col";
    col.innerHTML = `
      <div class="col-header">
        <div class="col-title">${{group}}</div>
        <div class="col-summary">
          <span class="si">${{inStk}} in stock</span>
          <span class="so">${{out}} out</span>
          <span class="pct">${{pct}}%</span>
        </div>
        <div class="prog-wrap"><div class="prog-bar" style="width:${{pct}}%;background:${{bc}}"></div></div>
      </div>
      ${{sectionsHTML}}`;
    wrap.appendChild(col);
  }});
}}

/* ── Main render ── */
function render(dateStr) {{
  const idx = DATA.dates.indexOf(dateStr);
  if (idx === -1) return;
  buildChart(idx);
  buildColumns(idx);
}}

/* ── Init ── */
populatePicker();
render(DATA.dates[DATA.dates.length - 1]);
</script>
</body>
</html>"""


if __name__ == "__main__":
    print("Reading:", XLSX_PATH)
    dates, filaments = read_spreadsheet(XLSX_PATH)
    print(f"Found {len(dates)} date(s):", dates)
    html = build_html(dates, filaments)
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w") as f:
        f.write(html)
    print("Generated:", OUT_PATH)
